import imaplib
import re
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import Font
from tqdm import tqdm

# connect to email server
mail = imaplib.IMAP4_SSL('imap.gmail.com')
#CHANGE YOUR EMAIL AND PASSWORD HERE
#THE PASSWORD NEEDS TO BE AN APP SPECIFIC PASSWORD
#MAKE SURE YOU HAVE 2FA TURNED ON AND GO TO THIS SITE
#https://myaccount.google.com/u/0/apppasswords?rapt=AEjHL4PoZPQ7LovSpkNjHn4Yp2UVlegzTah7sTWHVQM3wydBn8tDtINwbduObbIlvfolo3KzcE2v5qXqpviUv0g7wZTHfX5qDw&pageId=none&pli=1
mail.login('EMAIL', 'APP_SPECIFIC_PASS')
mail.select('inbox')

# search for emails
status, messages = mail.search(None, 'ALL')
emails = messages[0].split(b' ')

# create progress bar
pbar = tqdm(total=len(emails))

# create or load Excel workbook
try:
    wb = load_workbook('edu_emails.xlsx')
except FileNotFoundError:
    wb = Workbook()

# select the active worksheet
ws = wb.active

# add header to worksheet if it is empty
if not ws.cell(row=1, column=1).value:
    ws.cell(row=1, column=1).value = 'Email Address'
    ws.cell(row=1, column=2).value = 'Count'

# create dictionary to hold email counts
email_counts = {}

# loop through emails
for email in emails:
    # fetch email data
    status, data = mail.fetch(email, '(RFC822)')
    # extract email address
    email_address = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\.edu\b', str(data))
    # if email address is found, update count
    if email_address:
        email_address = email_address[0]
        if email_address in email_counts:
            email_counts[email_address] += 1
        else:
            email_counts[email_address] = 1
    # update progress bar
    pbar.update(1)

# write email addresses and counts to worksheet
for row, (email, count) in enumerate(email_counts.items(), start=2):
    ws.cell(row=row, column=1).value = email
    ws.cell(row=row, column=2).value = count

# Set the range to search for emails
start_row = 2
end_row = 84
start_col = column_index_from_string('A')
end_col = column_index_from_string('B')

# Write the title for the Sites column
sites_title = 'Sites'
sites_title_cell = ws.cell(row=1, column=end_col+1)
sites_title_cell.value = sites_title
sites_title_cell.font = Font(bold=True)

# Create a dictionary to store the unique domain names and their counts
domains = {}

# Iterate over the cells in the range with a progress bar
for row in tqdm(range(start_row, end_row+1), desc='Processing emails', unit=' row'):
    cell = ws.cell(row=row, column=start_col)
    email = cell.value

    # Extract the [name].edu part from the email using the last two parts
    parts = email.split('.')
    if len(parts) >= 2:
        domain_name = parts[-2] + '.' + parts[-1]

        # Check if the domain name already exists in the dictionary
        if domain_name in domains:
            # If it does, add the count to the existing domain name
            count = ws.cell(row=row, column=start_col+1).value
            domains[domain_name]['count'] += count
            domains[domain_name]['rows'].append(row)
        else:
            # If it doesn't, add the domain name and count to the dictionary
            count = ws.cell(row=row, column=start_col+1).value
            domains[domain_name] = {'count': count, 'rows': [row]}

            # Write the domain name as a hyperlink in column C
            hyperlink = f'https://{domain_name}'
            sites_cell = ws.cell(row=row, column=end_col+1)
            sites_cell.value = domain_name
            sites_cell.font = Font(underline='single', color='0563C1')
            sites_cell.hyperlink = hyperlink

# Remove duplicates by keeping the first occurrence of each domain name
for domain_name, info in domains.items():
    if len(info['rows']) > 1:
        # If there are duplicates, delete all but the first row
        for row in info['rows'][1:]:
            ws.delete_rows(row)

# Remove duplicates from the original A column
unique_emails = set()
for row in range(start_row, end_row+1):
    cell = ws.cell(row=row, column=start_col)
    email = cell.value
    if email in unique_emails:
        ws.delete_rows(row)
    else:
        unique_emails.add(email)

# Write the unique domain names and their counts to column B
for row in range(start_row, end_row+1):
    cell = ws.cell(row=row, column=start_col)
    email = cell.value
    parts = email.split('.')
    if len(parts) >= 2:
        domain_name = parts[-2] + '.' + parts[-1]
        if domain_name in domains:
            count = domains[domain_name]['count']
            count_cell = ws.cell(row=row, column=start_col+1)
            count_cell.value = count

# Save the updated workbook
wb.save('emails.xlsx')

# close progress bar
pbar.close()

# logout from email server
mail.close()
mail.logout()
